"""
Excel Export Module - VM to EC2 Mapping
Generates detailed Excel spreadsheet with VM analysis and AWS recommendations
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
from agents.config.config import output_folder_dir_path


def get_case_output_path(filename):
    """
    Get case-specific output path for Excel files
    
    Args:
        filename: Output filename
    
    Returns:
        Full path in case-specific output folder, or root output folder if case not available
    """
    try:
        from agents.core.case_output_manager import get_case_output_manager
        case_manager = get_case_output_manager()
        return case_manager.get_output_path(filename)
    except:
        # Fallback to root output folder
        return os.path.join(output_folder_dir_path, filename)

def create_vm_to_ec2_mapping_excel(detailed_results_df, backup_costs=None, output_filename='vm_to_ec2_mapping.xlsx'):
    """
    Create detailed Excel spreadsheet with VM to EC2 mapping
    
    Args:
        detailed_results_df: DataFrame from pricing calculator with VM and EC2 details
        backup_costs: Dict with backup cost details (optional)
        output_filename: Output Excel filename
    
    Returns:
        Path to generated Excel file
    """
    
    # Calculate per-VM backup costs if backup_costs provided
    vm_backup_costs = {}
    if backup_costs:
        # Distribute backup costs proportionally by storage
        total_storage = detailed_results_df['storage_gb'].sum()
        total_backup_cost = backup_costs.get('total_monthly', 0)
        
        for idx, row in detailed_results_df.iterrows():
            vm_name = row.get('vm_name', f'VM-{idx+1}')
            vm_storage = row.get('storage_gb', 0)
            
            # Proportional backup cost
            if total_storage > 0:
                vm_backup_cost = (vm_storage / total_storage) * total_backup_cost
            else:
                vm_backup_cost = 0
            
            vm_backup_costs[vm_name] = vm_backup_cost
    
    # Prepare data for Excel
    excel_data = []
    
    for idx, row in detailed_results_df.iterrows():
        vm_name = row.get('vm_name', f'VM-{idx+1}')
        backup_cost = vm_backup_costs.get(vm_name, 0)
        
        excel_row = {
            # On-Premises VM Details
            'VM Name': vm_name,
            'VM vCPU (Provisioned)': row.get('original_vcpu', row.get('vcpu', 0)),
            'VM Memory GB (Provisioned)': row.get('original_memory_gb', row.get('memory_gb', 0)),
            'VM Storage GB (Provisioned)': row.get('original_storage_gb', row.get('storage_gb', 0)),
            'VM OS': row.get('os', 'Unknown'),
            
            # Utilization Data (if available)
            'CPU Utilization %': row.get('cpu_util', 'N/A'),
            'Memory Utilization %': row.get('memory_util', 'N/A'),
            'Storage Used GB': row.get('storage_used_gb', 'N/A'),
            
            # Right-Sizing Applied
            'Right-Sizing Applied': 'Yes' if row.get('right_sizing_applied', False) else 'No',
            'vCPU Reduction %': row.get('vcpu_reduction', 0) if row.get('right_sizing_applied', False) else 0,
            'Memory Reduction %': row.get('memory_reduction', 0) if row.get('right_sizing_applied', False) else 0,
            'Storage Reduction %': row.get('storage_reduction', 0) if row.get('right_sizing_applied', False) else 0,
            
            # Right-Sized Specs (After Optimization)
            'Optimized vCPU': row.get('vcpu', 0),
            'Optimized Memory GB': row.get('memory_gb', 0),
            'Optimized Storage GB': row.get('storage_gb', 0),
            
            # AWS EC2 Recommendation
            'AWS Instance Type': row.get('instance_type', 'N/A'),
            'EC2 vCPU': _get_instance_vcpu(row.get('instance_type', '')),
            'EC2 Memory GB': _get_instance_memory(row.get('instance_type', '')),
            'EC2 OS Type': row.get('os_type', 'Linux'),
            
            # AWS Pricing (3-Year No Upfront RI)
            'EC2 Hourly Rate ($)': row.get('hourly_rate', 0),
            'EC2 Monthly Cost ($)': row.get('monthly_compute', 0),
            'EBS Storage GB': row.get('storage_gb', 0),
            'EBS Monthly Cost ($)': row.get('monthly_storage', 0),
            'Data Transfer Monthly ($)': row.get('monthly_data_transfer', 0),
            'Backup Monthly Cost ($)': backup_cost,
            'Total Monthly Cost ($)': row.get('monthly_total', 0) + backup_cost,
            'Total Annual Cost ($)': (row.get('monthly_total', 0) + backup_cost) * 12,
        }
        
        excel_data.append(excel_row)
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Generate Excel file with formatting
    output_path = get_case_output_path(output_filename)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='VM to EC2 Mapping', index=False)
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['VM to EC2 Mapping']
        
        # Apply formatting
        _format_worksheet(worksheet, len(df))
        
        # Add summary sheet
        _add_summary_sheet(workbook, df, backup_costs)
        
        # Add backup summary sheet if backup costs available
        if backup_costs:
            _add_backup_summary_sheet(workbook, backup_costs)
    
    print(f"✓ Excel export created: {output_path}")
    return output_path

def _get_instance_vcpu(instance_type):
    """Get vCPU count for instance type"""
    from agents.pricing.aws_pricing_calculator import AWSPricingCalculator
    calculator = AWSPricingCalculator()
    specs = calculator.INSTANCE_SPECS.get(instance_type, (0, 0))
    return specs[0]

def _get_instance_memory(instance_type):
    """Get memory GB for instance type"""
    from agents.pricing.aws_pricing_calculator import AWSPricingCalculator
    calculator = AWSPricingCalculator()
    specs = calculator.INSTANCE_SPECS.get(instance_type, (0, 0))
    return specs[1]

def _format_worksheet(worksheet, num_rows):
    """Apply formatting to worksheet"""
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    
    section_fills = {
        'vm': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'util': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'rightsizing': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
        'optimized': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'ec2': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'pricing': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
        'backup': PatternFill(start_color='C5E0B4', end_color='C5E0B4', fill_type='solid'),  # Light green for backup
    }
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths and apply section colors
    # Updated to include backup cost column (now 27 columns total)
    column_sections = [
        (1, 5, 'vm'),      # VM Name to VM OS
        (6, 8, 'util'),    # Utilization data
        (9, 12, 'rightsizing'),  # Right-sizing info
        (13, 15, 'optimized'),   # Optimized specs
        (16, 19, 'ec2'),   # EC2 details
        (20, 25, 'pricing'),  # Pricing details (EC2, EBS, Data Transfer)
        (26, 26, 'backup'),   # Backup cost
        (27, 28, 'pricing'),  # Total costs
    ]
    
    for col_start, col_end, section in column_sections:
        for col in range(col_start, col_end + 1):
            # Set column width
            worksheet.column_dimensions[worksheet.cell(1, col).column_letter].width = 18
            
            # Apply section color to data rows
            for row in range(2, num_rows + 2):
                cell = worksheet.cell(row, col)
                cell.fill = section_fills[section]
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Format numbers
                cell_value = worksheet.cell(1, col).value
                if cell_value and ('Cost' in cell_value or 'Rate' in cell_value):
                    cell.number_format = '$#,##0.00'
                elif cell_value and '%' in cell_value:
                    cell.number_format = '0.0'
                elif cell_value and ('GB' in cell_value or 'vCPU' in cell_value):
                    cell.number_format = '#,##0.0'
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions

def _add_summary_sheet(workbook, df, backup_costs=None):
    """Add summary sheet with aggregated statistics"""
    
    summary_sheet = workbook.create_sheet('Summary', 0)
    
    # Calculate summary statistics
    total_vms = len(df)
    total_monthly_cost = df['Total Monthly Cost ($)'].sum()
    total_annual_cost = df['Total Annual Cost ($)'].sum()
    
    # Backup costs
    total_backup_monthly = df['Backup Monthly Cost ($)'].sum() if 'Backup Monthly Cost ($)' in df.columns else 0
    total_backup_annual = total_backup_monthly * 12
    
    # Right-sizing statistics
    vms_right_sized = len(df[df['Right-Sizing Applied'] == 'Yes'])
    avg_vcpu_reduction = df[df['Right-Sizing Applied'] == 'Yes']['vCPU Reduction %'].mean() if vms_right_sized > 0 else 0
    avg_memory_reduction = df[df['Right-Sizing Applied'] == 'Yes']['Memory Reduction %'].mean() if vms_right_sized > 0 else 0
    avg_storage_reduction = df[df['Right-Sizing Applied'] == 'Yes']['Storage Reduction %'].mean() if vms_right_sized > 0 else 0
    
    # Instance type distribution
    instance_counts = df['AWS Instance Type'].value_counts()
    
    # OS distribution
    os_counts = df['EC2 OS Type'].value_counts()
    
    # Write summary data
    summary_data = [
        ['AWS Migration Cost Analysis - Summary'],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        [''],
        ['Overall Statistics'],
        ['Total VMs:', total_vms],
        ['Total Monthly Cost (incl. Backup):', f'${total_monthly_cost:,.2f}'],
        ['Total Annual Cost (ARR):', f'${total_annual_cost:,.2f}'],
        ['Average Cost per VM (Monthly):', f'${total_monthly_cost/total_vms:,.2f}' if total_vms > 0 else '$0'],
        [''],
        ['Cost Breakdown'],
        ['Compute & Storage Monthly:', f'${total_monthly_cost - total_backup_monthly:,.2f}'],
        ['Backup Monthly:', f'${total_backup_monthly:,.2f}'],
        ['Total Monthly:', f'${total_monthly_cost:,.2f}'],
        [''],
        ['Right-Sizing Statistics'],
        ['VMs Right-Sized:', vms_right_sized],
        ['VMs Not Right-Sized:', total_vms - vms_right_sized],
        ['Average vCPU Reduction:', f'{avg_vcpu_reduction:.1f}%'],
        ['Average Memory Reduction:', f'{avg_memory_reduction:.1f}%'],
        ['Average Storage Reduction:', f'{avg_storage_reduction:.1f}%'],
        [''],
        ['OS Distribution'],
    ]
    
    for os_type, count in os_counts.items():
        summary_data.append([f'{os_type} VMs:', count, f'({count/total_vms*100:.1f}%)'])
    
    summary_data.append([''])
    summary_data.append(['Top 10 Instance Types'])
    
    for instance_type, count in instance_counts.head(10).items():
        monthly_cost = df[df['AWS Instance Type'] == instance_type]['Total Monthly Cost ($)'].sum()
        summary_data.append([instance_type, f'{count} VMs', f'${monthly_cost:,.2f}/month'])
    
    # Write to sheet
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row_idx, col_idx, value)
            
            # Format headers
            if row_idx in [1, 4, 10, 15, 23]:  # Updated row numbers for headers
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            elif col_idx == 1 and row_idx > 1:
                cell.font = Font(bold=True)
    
    # Set column widths
    summary_sheet.column_dimensions['A'].width = 35
    summary_sheet.column_dimensions['B'].width = 20
    summary_sheet.column_dimensions['C'].width = 20

def _add_backup_summary_sheet(workbook, backup_costs):
    """Add backup summary sheet with detailed backup cost breakdown"""
    
    backup_sheet = workbook.create_sheet('Backup Summary')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, size=12, color='FFFFFF')
    section_font = Font(bold=True, size=11)
    
    # Title
    backup_sheet['A1'] = 'AWS Backup Cost Summary'
    backup_sheet['A1'].font = Font(bold=True, size=14)
    backup_sheet.merge_cells('A1:C1')
    
    backup_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    backup_sheet.merge_cells('A2:C2')
    
    # Overall backup costs
    row = 4
    backup_sheet[f'A{row}'] = 'Overall Backup Costs'
    backup_sheet[f'A{row}'].font = header_font
    backup_sheet[f'A{row}'].fill = header_fill
    backup_sheet.merge_cells(f'A{row}:C{row}')
    
    row += 1
    backup_sheet[f'A{row}'] = 'Total Monthly Cost:'
    backup_sheet[f'A{row}'].font = section_font
    backup_sheet[f'B{row}'] = f"${backup_costs.get('total_monthly', 0):,.2f}"
    
    row += 1
    backup_sheet[f'A{row}'] = 'Total Annual Cost:'
    backup_sheet[f'A{row}'].font = section_font
    backup_sheet[f'B{row}'] = f"${backup_costs.get('total_annual', 0):,.2f}"
    
    row += 1
    backup_sheet[f'A{row}'] = 'Storage Tiering Savings:'
    backup_sheet[f'A{row}'].font = section_font
    backup_sheet[f'B{row}'] = f"${backup_costs.get('savings_vs_all_warm', 0):,.2f}"
    backup_sheet[f'C{row}'] = f"({backup_costs.get('savings_percentage', 0):.1f}%)"
    
    # Production environment
    row += 2
    backup_sheet[f'A{row}'] = 'Production Environment'
    backup_sheet[f'A{row}'].font = header_font
    backup_sheet[f'A{row}'].fill = header_fill
    backup_sheet.merge_cells(f'A{row}:C{row}')
    
    prod = backup_costs.get('production', {})
    row += 1
    backup_sheet[f'A{row}'] = 'VM Count:'
    backup_sheet[f'B{row}'] = prod.get('vm_count', 0)
    
    row += 1
    backup_sheet[f'A{row}'] = 'Total Storage (GB):'
    backup_sheet[f'B{row}'] = f"{prod.get('storage_gb', 0):,.0f}"
    
    row += 1
    backup_sheet[f'A{row}'] = 'Backup Storage (GB):'
    backup_sheet[f'B{row}'] = f"{prod.get('backup_storage_gb', 0):,.0f}"
    backup_sheet[f'C{row}'] = '(after compression/dedup)'
    
    row += 1
    backup_sheet[f'A{row}'] = 'Monthly Cost:'
    backup_sheet[f'A{row}'].font = section_font
    backup_sheet[f'B{row}'] = f"${prod.get('monthly_cost', 0):,.2f}"
    
    row += 1
    backup_sheet[f'A{row}'] = 'Annual Cost:'
    backup_sheet[f'A{row}'].font = section_font
    backup_sheet[f'B{row}'] = f"${prod.get('annual_cost', 0):,.2f}"
    
    # Retention policy
    retention = prod.get('retention_summary', {})
    if retention:
        row += 1
        backup_sheet[f'A{row}'] = 'Retention Policy:'
        backup_sheet[f'A{row}'].font = section_font
        
        row += 1
        backup_sheet[f'A{row}'] = '  Daily:'
        backup_sheet[f'B{row}'] = retention.get('daily', 'N/A')
        
        row += 1
        backup_sheet[f'A{row}'] = '  Weekly:'
        backup_sheet[f'B{row}'] = retention.get('weekly', 'N/A')
        
        row += 1
        backup_sheet[f'A{row}'] = '  Monthly:'
        backup_sheet[f'B{row}'] = retention.get('monthly', 'N/A')
    
    # Storage tier breakdown
    tier_breakdown = prod.get('tier_breakdown', {})
    if tier_breakdown:
        row += 1
        backup_sheet[f'A{row}'] = 'Storage Tier Breakdown:'
        backup_sheet[f'A{row}'].font = section_font
        
        for tier_name, tier_data in tier_breakdown.items():
            if tier_data.get('gb', 0) > 0:
                row += 1
                backup_sheet[f'A{row}'] = f'  {tier_name.replace("_", " ").title()}:'
                backup_sheet[f'B{row}'] = f"{tier_data.get('gb', 0):,.0f} GB"
                backup_sheet[f'C{row}'] = f"${tier_data.get('cost', 0):,.2f}/month"
    
    # Pre-production environment
    row += 2
    backup_sheet[f'A{row}'] = 'Pre-Production Environment'
    backup_sheet[f'A{row}'].font = header_font
    backup_sheet[f'A{row}'].fill = header_fill
    backup_sheet.merge_cells(f'A{row}:C{row}')
    
    preprod = backup_costs.get('pre_production', {})
    row += 1
    backup_sheet[f'A{row}'] = 'VM Count:'
    backup_sheet[f'B{row}'] = preprod.get('vm_count', 0)
    
    row += 1
    backup_sheet[f'A{row}'] = 'Total Storage (GB):'
    backup_sheet[f'B{row}'] = f"{preprod.get('storage_gb', 0):,.0f}"
    
    row += 1
    backup_sheet[f'A{row}'] = 'Backup Storage (GB):'
    backup_sheet[f'B{row}'] = f"{preprod.get('backup_storage_gb', 0):,.0f}"
    backup_sheet[f'C{row}'] = '(after compression/dedup)'
    
    row += 1
    backup_sheet[f'A{row}'] = 'Monthly Cost:'
    backup_sheet[f'A{row}'].font = section_font
    backup_sheet[f'B{row}'] = f"${preprod.get('monthly_cost', 0):,.2f}"
    
    row += 1
    backup_sheet[f'A{row}'] = 'Annual Cost:'
    backup_sheet[f'A{row}'].font = section_font
    backup_sheet[f'B{row}'] = f"${preprod.get('annual_cost', 0):,.2f}"
    
    # Retention policy
    retention = preprod.get('retention_summary', {})
    if retention:
        row += 1
        backup_sheet[f'A{row}'] = 'Retention Policy:'
        backup_sheet[f'A{row}'].font = section_font
        
        row += 1
        backup_sheet[f'A{row}'] = '  Daily:'
        backup_sheet[f'B{row}'] = retention.get('daily', 'N/A')
        
        row += 1
        backup_sheet[f'A{row}'] = '  Weekly:'
        backup_sheet[f'B{row}'] = retention.get('weekly', 'N/A')
        
        row += 1
        backup_sheet[f'A{row}'] = '  Monthly:'
        backup_sheet[f'B{row}'] = retention.get('monthly', 'N/A')
    
    # Set column widths
    backup_sheet.column_dimensions['A'].width = 35
    backup_sheet.column_dimensions['B'].width = 20
    backup_sheet.column_dimensions['C'].width = 30


def export_vm_to_ec2_mapping(pricing_results, output_filename='vm_to_ec2_mapping.xlsx'):
    """
    Main export function - creates Excel with VM to EC2 mapping
    
    Args:
        pricing_results: Results dictionary from pricing calculator
        output_filename: Output filename
    
    Returns:
        Path to generated Excel file
    """
    detailed_df = pricing_results.get('detailed_results')
    backup_costs = pricing_results.get('backup_costs')
    
    if detailed_df is None or len(detailed_df) == 0:
        print("⚠ No detailed results available for Excel export")
        return None
    
    return create_vm_to_ec2_mapping_excel(detailed_df, backup_costs, output_filename)


if __name__ == "__main__":
    print("Excel Export Module - VM to EC2 Mapping")
    print("Use export_vm_to_ec2_mapping() to generate Excel reports")


def export_rvtools_dual_pricing(results_option1, results_option2, output_filename='vm_to_ec2_mapping.xlsx'):
    """
    Export RVTools pricing with BOTH pricing options to Excel
    
    Similar to IT Inventory dual pricing, but EC2-only (no RDS)
    - Option 1: 3-Year EC2 Instance Savings Plan
    - Option 2: 3-Year Compute Savings Plan
    
    Args:
        results_option1: Results from Option 1 (EC2 Instance SP)
        results_option2: Results from Option 2 (Compute SP)
        output_filename: Output Excel filename
    
    Returns:
        Path to generated Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import pandas as pd
    from datetime import datetime
    
    output_path = get_case_output_path(output_filename)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Tab 1: Pricing Comparison Summary
    monthly_savings = results_option2['summary']['total_monthly_cost'] - results_option1['summary']['total_monthly_cost']
    annual_savings = monthly_savings * 12
    three_year_savings = monthly_savings * 36
    savings_pct = (monthly_savings / results_option2['summary']['total_monthly_cost'] * 100) if results_option2['summary']['total_monthly_cost'] > 0 else 0
    
    # Get backup costs
    backup_monthly = results_option1['summary'].get('backup_monthly_cost', 0)
    backup_annual = results_option1['summary'].get('backup_annual_cost', 0)
    
    comparison_data = {
        'Metric': [
            'Total VMs',
            '',
            'Option 1: 3-Year EC2 Instance Savings Plan',
            'EC2 Monthly Cost (Compute + Storage)',
            'Backup Monthly Cost',
            'Total Monthly Cost (incl. Backup)',
            'Total Annual Cost (ARR)',
            '3-Year Pricing',
            '',
            'Option 2: 3-Year Compute Savings Plan',
            'EC2 Monthly Cost (Compute + Storage)',
            'Backup Monthly Cost',
            'Total Monthly Cost (incl. Backup)',
            'Total Annual Cost (ARR)',
            '3-Year Pricing',
            '',
            'Savings (Option 1 vs Option 2)',
            'Monthly Savings',
            'Annual Savings',
            '3-Year Savings',
            'Savings Percentage',
            '',
            'Region',
            'Recommendation'
        ],
        'Value': [
            results_option1['summary']['total_vms'],
            '',
            '',
            f"${results_option1['summary']['total_monthly_cost']:,.2f}",
            f"${backup_monthly:,.2f}",
            f"${results_option1['summary'].get('total_monthly_with_backup', results_option1['summary']['total_monthly_cost'] + backup_monthly):,.2f}",
            f"${results_option1['summary'].get('total_arr_with_backup', (results_option1['summary']['total_monthly_cost'] + backup_monthly) * 12):,.2f}",
            f"${results_option1['summary'].get('total_monthly_with_backup', results_option1['summary']['total_monthly_cost'] + backup_monthly) * 36:,.2f}",
            '',
            '',
            f"${results_option2['summary']['total_monthly_cost']:,.2f}",
            f"${backup_monthly:,.2f}",
            f"${results_option2['summary'].get('total_monthly_with_backup', results_option2['summary']['total_monthly_cost'] + backup_monthly):,.2f}",
            f"${results_option2['summary'].get('total_arr_with_backup', (results_option2['summary']['total_monthly_cost'] + backup_monthly) * 12):,.2f}",
            f"${results_option2['summary'].get('total_monthly_with_backup', results_option2['summary']['total_monthly_cost'] + backup_monthly) * 36:,.2f}",
            '',
            '',
            f"${monthly_savings:,.2f}",
            f"${annual_savings:,.2f}",
            f"${three_year_savings:,.2f}",
            f"{savings_pct:.2f}%",
            '',
            results_option1['summary']['region'],
            f"Option 1 saves ${monthly_savings:,.2f}/month ({savings_pct:.1f}%)"
        ]
    }
    
    df_comparison = pd.DataFrame(comparison_data)
    ws_comparison = wb.create_sheet('Pricing Comparison')
    
    # Write header row
    ws_comparison.cell(1, 1, 'Metric')
    ws_comparison.cell(1, 2, 'Value')
    ws_comparison['A1'].font = Font(bold=True, color='FFFFFF')
    ws_comparison['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws_comparison['B1'].font = Font(bold=True, color='FFFFFF')
    ws_comparison['B1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data rows (starting from row 2)
    for r_idx, row in enumerate(df_comparison.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_comparison.cell(r_idx, c_idx, value)
            
            # Format section headers (adjusted for header row and new backup rows)
            # Data indices: 2='Option 1', 9='Option 2', 16='Savings'
            # With header row: 2+2=4, 9+2=11, 16+2=18
            if r_idx in [4, 11, 18]:  # Section header rows
                cell.font = Font(bold=True, size=11, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            elif c_idx == 1 and r_idx > 1:  # Don't bold the header row again
                cell.font = Font(bold=True)
    
    ws_comparison.column_dimensions['A'].width = 40
    ws_comparison.column_dimensions['B'].width = 25
    
    # Tab 2: EC2 Details (Option 1 - EC2 Instance SP)
    ec2_details_option1 = []
    detailed_df_option1 = results_option1['detailed_results']
    
    for idx, row in detailed_df_option1.iterrows():
        from agents.pricing.aws_pricing_calculator import AWSPricingCalculator
        calculator = AWSPricingCalculator()
        instance_specs = calculator.INSTANCE_SPECS.get(row['instance_type'], (0, 0))
        
        ec2_details_option1.append({
            'VM Name': row['vm_name'],
            'VM vCPU': row.get('original_vcpu', row['vcpu']),
            'VM Memory GB': row.get('original_memory_gb', row['memory_gb']),
            'VM Storage GB': row.get('original_storage_gb', row['storage_gb']),
            'VM OS': row['os'],
            'Right-Sizing Applied': 'Yes' if row.get('right_sizing_applied', False) else 'No',
            'Optimized vCPU': row['vcpu'],
            'Optimized Memory GB': row['memory_gb'],
            'Optimized Storage GB': row['storage_gb'],
            'AWS Instance Type': row['instance_type'],
            'EC2 vCPU': instance_specs[0],
            'EC2 Memory GB': instance_specs[1],
            'EC2 OS Type': row['os_type'],
            'Pricing Model': '3-Year EC2 Instance SP',
            'EC2 Hourly Rate ($)': row['hourly_rate'],
            'EC2 Monthly Compute ($)': row['monthly_compute'],
            'EBS Monthly Storage ($)': row['monthly_storage'],
            'Data Transfer Monthly ($)': row['monthly_data_transfer'],
            'Total Monthly Cost ($)': row['monthly_total'],
            'Total Annual Cost ($)': row['monthly_total'] * 12
        })
    
    df_ec2_option1 = pd.DataFrame(ec2_details_option1)
    ws_ec2_option1 = wb.create_sheet('EC2 Details - Option 1')
    
    # Write headers first
    for c_idx, col_name in enumerate(df_ec2_option1.columns, 1):
        cell = ws_ec2_option1.cell(1, c_idx, col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data starting from row 2
    for r_idx, row in enumerate(df_ec2_option1.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_ec2_option1.cell(r_idx, c_idx, value)
    
    # Tab 3: EC2 Details (Option 2 - Compute SP)
    ec2_details_option2 = []
    detailed_df_option2 = results_option2['detailed_results']
    
    for idx, row in detailed_df_option2.iterrows():
        from agents.pricing.aws_pricing_calculator import AWSPricingCalculator
        calculator = AWSPricingCalculator()
        instance_specs = calculator.INSTANCE_SPECS.get(row['instance_type'], (0, 0))
        
        ec2_details_option2.append({
            'VM Name': row['vm_name'],
            'VM vCPU': row.get('original_vcpu', row['vcpu']),
            'VM Memory GB': row.get('original_memory_gb', row['memory_gb']),
            'VM Storage GB': row.get('original_storage_gb', row['storage_gb']),
            'VM OS': row['os'],
            'Right-Sizing Applied': 'Yes' if row.get('right_sizing_applied', False) else 'No',
            'Optimized vCPU': row['vcpu'],
            'Optimized Memory GB': row['memory_gb'],
            'Optimized Storage GB': row['storage_gb'],
            'AWS Instance Type': row['instance_type'],
            'EC2 vCPU': instance_specs[0],
            'EC2 Memory GB': instance_specs[1],
            'EC2 OS Type': row['os_type'],
            'Pricing Model': '3-Year Compute SP',
            'EC2 Hourly Rate ($)': row['hourly_rate'],
            'EC2 Monthly Compute ($)': row['monthly_compute'],
            'EBS Monthly Storage ($)': row['monthly_storage'],
            'Data Transfer Monthly ($)': row['monthly_data_transfer'],
            'Total Monthly Cost ($)': row['monthly_total'],
            'Total Annual Cost ($)': row['monthly_total'] * 12
        })
    
    df_ec2_option2 = pd.DataFrame(ec2_details_option2)
    ws_ec2_option2 = wb.create_sheet('EC2 Details - Option 2')
    
    # Write headers first
    for c_idx, col_name in enumerate(df_ec2_option2.columns, 1):
        cell = ws_ec2_option2.cell(1, c_idx, col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data starting from row 2
    for r_idx, row in enumerate(df_ec2_option2.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_ec2_option2.cell(r_idx, c_idx, value)
    
    # Tab 4: EC2 Comparison (Option 1 vs Option 2)
    ec2_comparison = []
    for idx, row_option1 in detailed_df_option1.iterrows():
        row_option2 = detailed_df_option2.iloc[idx]
        
        savings = row_option2['monthly_total'] - row_option1['monthly_total']
        savings_pct = (savings / row_option2['monthly_total'] * 100) if row_option2['monthly_total'] > 0 else 0
        
        ec2_comparison.append({
            'VM Name': row_option1['vm_name'],
            'Instance Type': row_option1['instance_type'],
            'OS Type': row_option1['os_type'],
            'Option 1 Monthly ($)': row_option1['monthly_total'],
            'Option 2 Monthly ($)': row_option2['monthly_total'],
            'Monthly Savings ($)': savings,
            'Savings %': f"{savings_pct:.2f}%"
        })
    
    df_ec2_comparison = pd.DataFrame(ec2_comparison)
    ws_ec2_comparison = wb.create_sheet('EC2 Comparison')
    
    # Write headers first
    for c_idx, col_name in enumerate(df_ec2_comparison.columns, 1):
        cell = ws_ec2_comparison.cell(1, c_idx, col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Write data starting from row 2
    for r_idx, row in enumerate(df_ec2_comparison.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws_ec2_comparison.cell(r_idx, c_idx, value)
    
    # Tab 5: Backup Summary (if backup costs available)
    backup_costs = results_option1.get('backup_costs', {})
    if backup_costs and backup_costs.get('total_monthly', 0) > 0:
        _add_backup_summary_sheet(wb, backup_costs)
        print(f"✓ Added Backup Summary tab with ${backup_costs.get('total_monthly', 0):,.2f}/month")
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Excel export created: {output_path}")
    
    return output_path


# ============================================================================
# EKS EXCEL EXPORT FUNCTIONS (Phase 4)
# ============================================================================

def create_eks_vm_details_tab(workbook, eks_vms, ec2_costs_per_vm, eks_costs_per_vm, region='us-east-1'):
    """
    Create EKS VM Details tab - similar to EC2 Excel with per-VM breakdown
    
    Shows which 16 VMs are migrated to EKS with individual EC2 vs EKS costs
    
    Args:
        workbook: openpyxl Workbook object
        eks_vms: List of VMs being migrated to EKS
        ec2_costs_per_vm: Dict mapping VM name to EC2 cost
        eks_costs_per_vm: Dict mapping VM name to EKS cost
        region: AWS region
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from agents.pricing.aws_pricing_calculator import AWSPricingCalculator
    
    ws = workbook.create_sheet('EKS VM Details')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f'EKS VM Migration Details - {len(eks_vms)} VMs'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:M1')
    
    ws['A2'] = 'Individual VM costs for EC2-only vs EKS migration'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:M2')
    
    ws['A3'] = 'NOTE: EC2 costs shown are for these VMs if they stayed on EC2 (not EKS worker node infrastructure)'
    ws['A3'].font = Font(italic=True, size=9, color='FF0000')
    ws.merge_cells('A3:M3')
    
    # Headers
    headers = [
        'VM Name', 'vCPU', 'Memory (GB)', 'Storage (GB)', 'OS',
        'EC2 Instance Type', 'EC2 Monthly ($)', 
        'EKS Share Monthly ($)', 'Difference ($)', 'Difference %',
        'Migration Target', 'Consolidation Ratio', 'Notes'
    ]
    
    row = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Get calculator for instance type mapping
    calculator = AWSPricingCalculator(region=region)
    
    # Data rows
    row = 6
    total_ec2 = 0
    total_eks = 0
    
    for vm in eks_vms:
        vm_name = vm.get('name', vm.get('vm_name', f'VM-{row-4}'))
        vcpu = vm.get('vcpu', 0)
        memory_gb = vm.get('memory', vm.get('memory_gb', 0))
        storage_gb = vm.get('storage_gb', 100)
        os_type = vm.get('os', 'Linux')
        
        # Get EC2 instance recommendation
        instance_type = calculator.map_vm_to_instance_type(vcpu, memory_gb, 'Windows' if 'windows' in os_type.lower() else 'Linux')
        
        # Get costs
        ec2_monthly = ec2_costs_per_vm.get(vm_name, 0)
        eks_monthly = eks_costs_per_vm.get(vm_name, 0)
        difference = eks_monthly - ec2_monthly
        difference_pct = (difference / ec2_monthly * 100) if ec2_monthly > 0 else 0
        
        # Consolidation ratio (from VM categorization)
        consolidation_ratio = vm.get('consolidation_ratio', 3.0)
        
        # Notes
        notes = vm.get('migration_reason', 'Containerizable workload')
        
        # Write row
        ws.cell(row, 1, vm_name)
        ws.cell(row, 2, vcpu)
        ws.cell(row, 3, memory_gb)
        ws.cell(row, 4, storage_gb)
        ws.cell(row, 5, os_type)
        ws.cell(row, 6, instance_type)
        ws.cell(row, 7, ec2_monthly).number_format = '$#,##0.00'
        ws.cell(row, 8, eks_monthly).number_format = '$#,##0.00'
        ws.cell(row, 9, difference).number_format = '$#,##0.00'
        ws.cell(row, 10, difference_pct).number_format = '0.0"%"'
        ws.cell(row, 11, 'EKS')
        ws.cell(row, 12, f'{consolidation_ratio:.2f}x')
        ws.cell(row, 13, notes)
        
        # Apply borders
        for col in range(1, 14):
            ws.cell(row, col).border = border
        
        total_ec2 += ec2_monthly
        total_eks += eks_monthly
        row += 1
    
    # Totals row
    ws.cell(row, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(row, 7, total_ec2).number_format = '$#,##0.00'
    ws.cell(row, 7).font = Font(bold=True)
    ws.cell(row, 8, total_eks).number_format = '$#,##0.00'
    ws.cell(row, 8).font = Font(bold=True)
    ws.cell(row, 9, total_eks - total_ec2).number_format = '$#,##0.00'
    ws.cell(row, 9).font = Font(bold=True)
    
    for col in range(1, 14):
        ws.cell(row, col).border = border
    
    # Add explanation row
    row += 2
    ws.cell(row, 1, 'IMPORTANT: The EC2 costs above are what these VMs would cost if they stayed on EC2.')
    ws.cell(row, 1).font = Font(italic=True, size=10, color='FF0000')
    ws.merge_cells(f'A{row}:M{row}')
    
    row += 1
    ws.cell(row, 1, 'EKS worker node infrastructure costs are shown separately in the EKS Cluster Design tab.')
    ws.cell(row, 1).font = Font(italic=True, size=10, color='FF0000')
    ws.merge_cells(f'A{row}:M{row}')
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 30
    
    return ws


def create_eks_sizing_tab(workbook, vm_categorization, strategy='hybrid'):
    """
    Create EKS VM Sizing & Categorization tab
    
    Shows which VMs go to EKS vs EC2 based on OS and strategy
    
    Args:
        workbook: openpyxl Workbook object
        vm_categorization: Dict from categorize_vms_for_eks()
        strategy: 'hybrid', 'all-eks', or 'disabled'
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = workbook.create_sheet('EKS VM Categorization')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f'EKS VM Categorization - Strategy: {strategy.upper()}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Summary section
    ws['A3'] = 'Category'
    ws['B3'] = 'VM Count'
    ws['C3'] = 'Total vCPU'
    ws['D3'] = 'Total Memory (GB)'
    ws['E3'] = 'Avg vCPU/VM'
    ws['F3'] = 'Destination'
    
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Data rows
    # NOTE: Only show VMs that are actually going to EKS in this tab
    # Database VMs go to EC2/RDS, so exclude them from EKS categorization
    categories = [
        ('Linux Small VMs', vm_categorization.get('linux_small', []), 'EKS'),
        ('Linux Medium VMs', vm_categorization.get('linux_medium', []), 'EKS'),
        ('Linux Large VMs', vm_categorization.get('linux_large', []), 'EKS'),
        ('Linux XLarge VMs', vm_categorization.get('linux_xlarge', []), 'EKS'),
        ('Windows Small VMs', vm_categorization.get('windows_small', []), 'EC2' if strategy == 'hybrid' else 'EKS'),
        ('Windows Medium VMs', vm_categorization.get('windows_medium', []), 'EC2' if strategy == 'hybrid' else 'EKS'),
        ('Windows Large VMs', vm_categorization.get('windows_large', []), 'EC2' if strategy == 'hybrid' else 'EKS'),
        ('Windows XLarge VMs', vm_categorization.get('windows_xlarge', []), 'EC2' if strategy == 'hybrid' else 'EKS'),
        # Database VMs excluded - they go to EC2/RDS, not EKS
        # ('Database VMs', vm_categorization.get('databases', []), 'EC2/RDS'),
    ]
    
    row = 4
    for category_name, vms, destination in categories:
        if not vms:
            continue
        
        vm_count = len(vms)
        total_vcpu = sum(vm.get('vcpu', 0) for vm in vms)
        total_memory = sum(vm.get('memory', vm.get('memory_gb', 0)) for vm in vms)
        avg_vcpu = total_vcpu / vm_count if vm_count > 0 else 0
        
        ws[f'A{row}'] = category_name
        ws[f'B{row}'] = vm_count
        ws[f'C{row}'] = total_vcpu
        ws[f'D{row}'] = round(total_memory, 1)
        ws[f'E{row}'] = round(avg_vcpu, 1)
        ws[f'F{row}'] = destination
        
        # Apply border
        for col in range(1, 7):
            ws.cell(row, col).border = border
        
        row += 1
    
    # Totals row
    total_vms = sum(len(vms) for _, vms, _ in categories if vms)
    total_vcpu = sum(sum(vm.get('vcpu', 0) for vm in vms) for _, vms, _ in categories if vms)
    total_memory = sum(sum(vm.get('memory', vm.get('memory_gb', 0)) for vm in vms) for _, vms, _ in categories if vms)
    
    ws[f'A{row}'] = 'TOTAL'
    ws[f'B{row}'] = total_vms
    ws[f'C{row}'] = total_vcpu
    ws[f'D{row}'] = round(total_memory, 1)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'].font = Font(bold=True)
    
    # Apply border to totals row
    for col in range(1, 7):
        ws.cell(row, col).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    return ws


def create_eks_capacity_mix_tab(workbook, cluster_config, eks_costs, region='us-east-1'):
    """
    Create EKS Capacity Mix Breakdown tab
    
    Shows detailed breakdown of Reserved/On-Demand/Spot capacity mix
    
    Args:
        workbook: openpyxl Workbook object
        cluster_config: Dict from calculate_eks_cluster_size()
        eks_costs: Dict with EKS pricing
        region: AWS region
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from agents.config.config import EKS_CONFIG
    
    ws = workbook.create_sheet('EKS Capacity Mix')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'EKS Capacity Mix Breakdown'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = 'Detailed breakdown of Reserved, On-Demand, and Spot instance costs'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Configuration section
    row = 4
    ws[f'A{row}'] = 'Capacity Mix Configuration'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = 'Capacity Type'
    ws[f'B{row}'] = 'Percentage'
    ws[f'C{row}'] = 'Pricing Model'
    ws[f'D{row}'] = 'Discount vs On-Demand'
    ws[f'E{row}'] = 'Configurable In'
    
    for col in range(1, 6):
        ws.cell(row, col).fill = header_fill
        ws.cell(row, col).font = header_font
        ws.cell(row, col).border = border
    
    row += 1
    
    # Get configuration
    reserved_pct = EKS_CONFIG.get('reserved_percentage', 0.30)
    on_demand_pct = EKS_CONFIG.get('on_demand_percentage', 0.40)
    spot_pct = EKS_CONFIG.get('spot_percentage', 0.30)
    spot_discount = EKS_CONFIG.get('spot_discount', 0.70)
    
    capacity_config = [
        ('Reserved Instances', f'{reserved_pct*100:.0f}%', '3-Year RI No Upfront', '40% discount', 'config.py: reserved_percentage'),
        ('On-Demand', f'{on_demand_pct*100:.0f}%', 'On-Demand pricing', 'Baseline (0%)', 'config.py: on_demand_percentage'),
        ('Spot Instances', f'{spot_pct*100:.0f}%', 'Spot pricing', f'{spot_discount*100:.0f}% discount', 'config.py: spot_percentage'),
    ]
    
    for capacity_type, percentage, pricing_model, discount, config_location in capacity_config:
        ws.cell(row, 1, capacity_type)
        ws.cell(row, 2, percentage)
        ws.cell(row, 3, pricing_model)
        ws.cell(row, 4, discount)
        ws.cell(row, 5, config_location)
        
        for col in range(1, 6):
            ws.cell(row, col).border = border
        row += 1
    
    # Cost breakdown section
    row += 1
    ws[f'A{row}'] = 'Cost Breakdown by Capacity Type'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = 'Capacity Type'
    ws[f'B{row}'] = 'Hourly Cost ($)'
    ws[f'C{row}'] = 'Monthly Cost ($)'
    ws[f'D{row}'] = 'Annual Cost ($)'
    ws[f'E{row}'] = '3-Year Cost ($)'
    
    for col in range(1, 6):
        ws.cell(row, col).fill = header_fill
        ws.cell(row, col).font = header_font
        ws.cell(row, col).border = border
    
    row += 1
    
    # Calculate costs from worker nodes
    worker_nodes = cluster_config.get('all_worker_nodes', [])
    total_hourly_ri = sum(node['hourly_rate'] * node.get('count', 1) for node in worker_nodes)
    
    # Calculate each capacity type
    reserved_hourly = total_hourly_ri * reserved_pct
    on_demand_hourly = total_hourly_ri * on_demand_pct * 1.67  # OD is ~1.67x RI
    spot_hourly = total_hourly_ri * spot_pct * 1.67 * (1 - spot_discount)  # Spot discount from OD
    combined_hourly = reserved_hourly + on_demand_hourly + spot_hourly
    
    cost_breakdown = [
        ('Reserved (30%)', reserved_hourly, reserved_hourly * 730, reserved_hourly * 730 * 12, reserved_hourly * 730 * 36),
        ('On-Demand (40% at 1.67x)', on_demand_hourly, on_demand_hourly * 730, on_demand_hourly * 730 * 12, on_demand_hourly * 730 * 36),
        ('Spot (30% at 70% discount)', spot_hourly, spot_hourly * 730, spot_hourly * 730 * 12, spot_hourly * 730 * 36),
        ('COMBINED TOTAL', combined_hourly, combined_hourly * 730, combined_hourly * 730 * 12, combined_hourly * 730 * 36),
    ]
    
    for capacity_type, hourly, monthly, annual, three_year in cost_breakdown:
        ws.cell(row, 1, capacity_type)
        ws.cell(row, 2, hourly).number_format = '$#,##0.0000'
        ws.cell(row, 3, monthly).number_format = '$#,##0.00'
        ws.cell(row, 4, annual).number_format = '$#,##0.00'
        ws.cell(row, 5, three_year).number_format = '$#,##0.00'
        
        if 'TOTAL' in capacity_type:
            for col in range(1, 6):
                ws.cell(row, col).font = Font(bold=True)
        
        for col in range(1, 6):
            ws.cell(row, col).border = border
        row += 1
    
    # Savings calculation
    row += 1
    ws[f'A{row}'] = 'Savings Analysis'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    
    # Calculate savings vs all On-Demand
    all_on_demand_hourly = total_hourly_ri * 1.67
    all_on_demand_monthly = all_on_demand_hourly * 730
    actual_monthly = combined_hourly * 730
    savings_monthly = all_on_demand_monthly - actual_monthly
    savings_pct = (savings_monthly / all_on_demand_monthly * 100) if all_on_demand_monthly > 0 else 0
    
    savings_data = [
        ('All On-Demand Cost (baseline)', '', all_on_demand_monthly, all_on_demand_monthly * 12, all_on_demand_monthly * 36),
        ('Actual Cost (with mix)', '', actual_monthly, actual_monthly * 12, actual_monthly * 36),
        ('Monthly Savings', '', savings_monthly, savings_monthly * 12, savings_monthly * 36),
        ('Savings Percentage', f'{savings_pct:.1f}%', '', '', ''),
    ]
    
    for label, hourly, monthly, annual, three_year in savings_data:
        ws.cell(row, 1, label)
        ws.cell(row, 2, hourly)
        if monthly:
            ws.cell(row, 3, monthly).number_format = '$#,##0.00'
        if annual:
            ws.cell(row, 4, annual).number_format = '$#,##0.00'
        if three_year:
            ws.cell(row, 5, three_year).number_format = '$#,##0.00'
        
        if 'Savings' in label:
            for col in range(1, 6):
                ws.cell(row, col).font = Font(bold=True)
                ws.cell(row, col).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        for col in range(1, 6):
            ws.cell(row, col).border = border
        row += 1
    
    # Configuration notes
    row += 2
    ws[f'A{row}'] = 'Configuration Notes:'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    ws[f'A{row}'] = f'• Capacity mix percentages are configurable in agents/config.py (EKS_CONFIG section)'
    row += 1
    ws[f'A{row}'] = f'• Reserved: {reserved_pct*100:.0f}% at 3-Year RI No Upfront rate (40% discount vs On-Demand)'
    row += 1
    ws[f'A{row}'] = f'• On-Demand: {on_demand_pct*100:.0f}% at baseline On-Demand rate (~1.67x Reserved rate)'
    row += 1
    ws[f'A{row}'] = f'• Spot: {spot_pct*100:.0f}% at {spot_discount*100:.0f}% discount from On-Demand (best for burst capacity)'
    row += 1
    ws[f'A{row}'] = f'• Total savings: {savings_pct:.1f}% vs all On-Demand pricing'
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 30
    
    return ws


def create_eks_cluster_design_tab(workbook, cluster_config, region='us-east-1'):
    """
    Create EKS Cluster Design tab
    
    Shows worker node configuration and cluster sizing
    
    Args:
        workbook: openpyxl Workbook object
        cluster_config: Dict from calculate_eks_cluster_size()
        region: AWS region
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = workbook.create_sheet('EKS Cluster Design')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f'EKS Cluster Design - Region: {region}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Cluster Overview
    ws['A3'] = 'Cluster Configuration'
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = section_fill
    ws.merge_cells('A3:E3')
    
    overview_data = [
        ('Total Worker Nodes', cluster_config.get('total_nodes', 0)),
        ('Total vCPU Capacity', cluster_config.get('total_vcpu', 0)),
        ('Total Memory (GB)', round(cluster_config.get('total_memory_gb', 0), 1)),
        ('Consolidation Ratio', f"{cluster_config.get('consolidation_ratio', 0):.2f}x"),
        ('Original VM vCPU', cluster_config.get('original_vcpu', 0)),
    ]
    
    row = 4
    for label, value in overview_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Worker Node Details
    row += 1
    ws[f'A{row}'] = 'Worker Node Configuration'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    headers = ['Instance Type', 'Count', 'vCPU/Node', 'Memory/Node (GB)', 'Purchase Option']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    row += 1
    worker_nodes = cluster_config.get('worker_nodes', [])
    if not worker_nodes:
        # Fallback: use all_worker_nodes if worker_nodes is empty
        worker_nodes = cluster_config.get('all_worker_nodes', [])
    
    for node in worker_nodes:
        instance_type = node.get('instance_type', 'N/A')
        ws.cell(row, 1, instance_type)
        ws.cell(row, 2, node.get('count', 0))
        ws.cell(row, 3, node.get('vcpu', 0))
        ws.cell(row, 4, round(node.get('memory_gb', 0), 1))
        ws.cell(row, 5, '3-Year Reserved Instance')  # Changed from On-Demand to 3yr RI
        
        for col in range(1, 6):
            ws.cell(row, col).border = border
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    
    return ws


def create_ec2_vs_eks_comparison_tab(workbook, ec2_costs, eks_costs, vm_categorization, cluster_config=None):
    """
    Create EC2 vs EKS Cost Comparison tab
    
    Shows side-by-side cost comparison for VMs planned for EKS migration
    
    Args:
        workbook: openpyxl Workbook object
        ec2_costs: Dict with EC2 pricing (for EKS VMs only)
        eks_costs: Dict with EKS pricing
        vm_categorization: Dict from categorize_vms_for_eks()
        cluster_config: Optional dict from calculate_eks_cluster_size() for detailed notes
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = workbook.create_sheet('EC2 vs EKS Comparison')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    savings_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    cost_increase_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'EC2 vs EKS Cost Comparison (EKS VMs Only)'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Get EKS VM count
    summary = vm_categorization.get('summary', {})
    eks_vm_count = summary.get('eks_linux', 0) + summary.get('eks_windows', 0)
    
    ws['A2'] = f'Comparing costs for {eks_vm_count} VMs planned for EKS migration'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:D2')
    
    ws['A3'] = 'All costs shown are MONTHLY rates based on 3-Year Reserved Instance pricing'
    ws['A3'].font = Font(italic=True, size=9, color='666666')
    ws.merge_cells('A3:D3')
    
    # Headers
    ws['A5'] = 'Cost Component'
    ws['B5'] = 'EC2-Only'
    ws['C5'] = 'EKS Hybrid'
    ws['D5'] = 'Difference'
    
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Cost breakdown - use actual costs passed in
    ec2_compute = ec2_costs.get('compute_monthly', 0)
    ec2_storage = ec2_costs.get('storage_monthly', 0)
    
    eks_compute = eks_costs.get('compute_monthly', 0)
    eks_storage = eks_costs.get('storage_monthly', 0)
    eks_control_plane = eks_costs.get('control_plane_monthly', 0)
    eks_networking = eks_costs.get('networking_monthly', 0)
    eks_monitoring = eks_costs.get('monitoring_monthly', 0)
    
    # Equal costs for fair comparison
    # Note: Networking cost (10% of compute) already includes data transfer, ALB, NAT Gateway
    ec2_networking_monthly = eks_networking  # Use same networking cost (10% of compute)
    ec2_monitoring_monthly = eks_monitoring  # Use same monitoring cost for fair comparison
    
    ec2_monthly = ec2_compute + ec2_storage + ec2_networking_monthly + ec2_monitoring_monthly
    
    # DEBUG: Log the control plane cost
    import logging
    logger = logging.getLogger('AgentWorkflow')
    logger.info(f"Excel Export - Control Plane Monthly: ${eks_control_plane:,.2f}")
    logger.info(f"Excel Export - EKS Costs Keys: {list(eks_costs.keys())}")
    
    eks_monthly = eks_compute + eks_storage + eks_control_plane + eks_networking + eks_monitoring
    
    difference = eks_monthly - ec2_monthly  # Positive = EKS costs more, Negative = EKS saves
    difference_pct = (difference / ec2_monthly * 100) if ec2_monthly > 0 else 0
    
    cost_data = [
        ('Compute Costs', ec2_compute, eks_compute),
        ('Storage Costs (EBS gp3)', ec2_storage, eks_storage),
        ('Networking (includes data transfer)', ec2_networking_monthly, eks_networking),
        ('EKS Control Plane', 0, eks_control_plane),
        ('Monitoring (CloudWatch)', ec2_monitoring_monthly, eks_monitoring),
        ('TOTAL MONTHLY', ec2_monthly, eks_monthly),
    ]
    
    row = 6
    for component, ec2_cost, eks_cost in cost_data:
        ws[f'A{row}'] = component
        ws[f'B{row}'] = ec2_cost
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'] = eks_cost
        ws[f'C{row}'].number_format = '$#,##0.00'
        
        if component == 'TOTAL MONTHLY':
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'D{row}'] = f'${difference:,.2f} ({difference_pct:+.1f}%)'
            # Use green for savings (negative difference), red for cost increase (positive difference)
            ws[f'D{row}'].fill = savings_fill if difference < 0 else cost_increase_fill
            ws[f'D{row}'].font = Font(bold=True)
        
        for col in range(1, 5):
            ws.cell(row, col).border = border
        row += 1
    
    # Annual costs
    row += 1
    ws[f'A{row}'] = 'TOTAL ANNUAL (ARR)'
    ws[f'B{row}'] = ec2_monthly * 12
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'] = eks_monthly * 12
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'] = f'${difference * 12:,.2f}'
    ws[f'D{row}'].fill = savings_fill if difference < 0 else cost_increase_fill
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'].font = Font(bold=True)
    
    for col in range(1, 5):
        ws.cell(row, col).border = border
    
    # 3-Year costs
    row += 1
    ws[f'A{row}'] = 'TOTAL 3-YEAR'
    ws[f'B{row}'] = ec2_monthly * 36
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'C{row}'] = eks_monthly * 36
    ws[f'C{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'] = f'${difference * 36:,.2f}'
    ws[f'D{row}'].fill = savings_fill if difference < 0 else cost_increase_fill
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'].font = Font(bold=True)
    
    for col in range(1, 5):
        ws.cell(row, col).border = border
    
    # Add pricing notes
    row += 2
    ws[f'A{row}'] = 'PRICING NOTES:'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    ws[f'A{row}'] = f'• Comparison is for {eks_vm_count} VMs planned for EKS migration only'
    row += 1
    ws[f'A{row}'] = '• EC2-Only: Costs if these VMs migrated to EC2 with 3-Year Reserved Instances'
    row += 1
    ws[f'A{row}'] = '• EKS Hybrid: Actual costs for containerizing these VMs on EKS with 3-Year Reserved Instances'
    row += 1
    ws[f'A{row}'] = '• Networking: Simplified to 10% of compute cost (includes data transfer, ALB, NAT Gateway, and data processing)'
    row += 1
    ws[f'A{row}'] = '• Storage: EC2 and EKS may differ due to different storage requirements (EKS uses container-optimized volumes)'
    row += 1
    
    # Add detailed monitoring cost calculation
    if cluster_config:
        num_nodes = cluster_config.get('total_nodes', 0)
    else:
        # Fallback: calculate from monitoring cost
        monitoring_monthly = eks_costs.get('monitoring_monthly', 0)
        num_nodes = int(monitoring_monthly / (60 * 0.50)) if monitoring_monthly > 0 else 0
    
    if num_nodes > 0:
        ws[f'A{row}'] = f'• CloudWatch Monitoring: EKS = {num_nodes} nodes × 60GB/month × $0.50/GB = ${eks_monitoring:,.2f}/month'
        row += 1
        ws[f'A{row}'] = '  EC2 baseline monitoring = $30/month (less detailed metrics)'
    else:
        ws[f'A{row}'] = '• CloudWatch: Both need monitoring, EKS requires more for container logs and metrics'
    row += 1
    ws[f'A{row}'] = '• Positive difference = EKS costs more (strategic investment for modernization)'
    row += 1
    ws[f'A{row}'] = '• Negative difference = EKS saves money (consolidation efficiency)'
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    
    return ws


def create_hybrid_migration_summary_tab(workbook, vm_categorization, ec2_costs_all_vms, eks_costs, backup_costs=None, region='us-east-1'):
    """
    Create Hybrid Migration Summary tab
    
    Shows complete hybrid migration cost with:
    - All EC2 VMs listed individually (VMs staying on EC2)
    - EKS infrastructure cost as one line item
    - Backup costs (EC2 + EKS)
    - Total hybrid migration cost
    
    Args:
        workbook: openpyxl Workbook object
        vm_categorization: Dict from categorize_vms_for_eks()
        ec2_costs_all_vms: Dict mapping VM name to EC2 cost for ALL VMs
        eks_costs: Dict with EKS pricing
        backup_costs: Optional dict with backup costs
        region: AWS region
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = workbook.create_sheet('Hybrid Migration Summary', 0)  # Insert as first tab
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    eks_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'Hybrid Migration Cost Summary'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f'Complete cost breakdown for hybrid EC2 + EKS migration strategy'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:D2')
    
    # Get EC2 and EKS VMs
    ec2_vms = vm_categorization.get('ec2', [])
    eks_vms = vm_categorization.get('eks_linux', []) + vm_categorization.get('eks_windows', [])
    
    # Section 1: EC2 VMs (Direct Migration)
    row = 4
    ws[f'A{row}'] = 'EC2 VMs (Direct Migration to EC2)'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'VM Name'
    ws[f'B{row}'] = 'vCPU'
    ws[f'C{row}'] = 'Memory (GB)'
    ws[f'D{row}'] = 'Monthly Cost ($)'
    
    for col in range(1, 5):
        ws.cell(row, col).fill = header_fill
        ws.cell(row, col).font = header_font
        ws.cell(row, col).border = border
    
    row += 1
    ec2_total = 0
    
    for vm in ec2_vms:
        vm_name = vm.get('name', vm.get('vm_name', 'Unknown'))
        vcpu = vm.get('vcpu', 0)
        memory = vm.get('memory', vm.get('memory_gb', 0))
        
        # Get EC2 cost from the mapping
        ec2_cost = ec2_costs_all_vms.get(vm_name, 0)
        
        ws[f'A{row}'] = vm_name
        ws[f'B{row}'] = vcpu
        ws[f'C{row}'] = memory
        ws[f'D{row}'] = ec2_cost
        ws[f'D{row}'].number_format = '$#,##0.00'
        
        for col in range(1, 5):
            ws.cell(row, col).border = border
        
        ec2_total += ec2_cost
        row += 1
    
    # EC2 Subtotal
    ws[f'A{row}'] = f'EC2 Subtotal ({len(ec2_vms)} VMs)'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = ec2_total
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    for col in range(1, 5):
        ws.cell(row, col).border = border
        ws.cell(row, col).fill = section_fill
    
    # Section 2: EKS Infrastructure
    row += 2
    ws[f'A{row}'] = 'EKS Infrastructure (Containerized Workloads)'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Component'
    ws[f'B{row}'] = 'Details'
    ws[f'C{row}'] = ''
    ws[f'D{row}'] = 'Monthly Cost ($)'
    
    for col in range(1, 5):
        ws.cell(row, col).fill = header_fill
        ws.cell(row, col).font = header_font
        ws.cell(row, col).border = border
    
    row += 1
    
    # EKS cost components
    eks_components = [
        ('EKS Control Plane', f'Managed Kubernetes control plane', eks_costs.get('control_plane_monthly', 0)),
        ('Worker Node Compute', f'{len(eks_vms)} VMs containerized', eks_costs.get('compute_monthly', 0)),
        ('Worker Node Storage', 'EBS gp3 volumes', eks_costs.get('storage_monthly', 0)),
        ('Networking', 'ALB, NAT Gateway, data transfer', eks_costs.get('networking_monthly', 0)),
        ('Monitoring', 'CloudWatch logs and metrics', eks_costs.get('monitoring_monthly', 0)),
    ]
    
    eks_total = 0
    for component, details, cost in eks_components:
        ws[f'A{row}'] = component
        ws[f'B{row}'] = details
        ws[f'D{row}'] = cost
        ws[f'D{row}'].number_format = '$#,##0.00'
        
        for col in range(1, 5):
            ws.cell(row, col).border = border
            ws.cell(row, col).fill = eks_fill
        
        eks_total += cost
        row += 1
    
    # EKS Subtotal
    ws[f'A{row}'] = f'EKS Subtotal ({len(eks_vms)} VMs containerized)'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = eks_total
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    for col in range(1, 5):
        ws.cell(row, col).border = border
        ws.cell(row, col).fill = section_fill
    
    # Section 3: Backup Costs
    row += 2
    backup_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    ws[f'A{row}'] = 'Backup Costs'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = 'Component'
    ws[f'B{row}'] = 'Details'
    ws[f'C{row}'] = ''
    ws[f'D{row}'] = 'Monthly Cost ($)'
    
    for col in range(1, 5):
        ws.cell(row, col).fill = header_fill
        ws.cell(row, col).font = header_font
        ws.cell(row, col).border = border
    
    row += 1
    
    # Calculate backup costs
    backup_monthly = 0
    ec2_backup_monthly = 0
    eks_backup_monthly = 0
    
    if backup_costs:
        # Get total backup cost
        backup_monthly = backup_costs.get('total_monthly', 0)
        
        # Try to get EC2 and EKS backup split if available
        ec2_backup_monthly = backup_costs.get('ec2_backup_monthly', 0)
        eks_backup_monthly = backup_costs.get('eks_backup_monthly', 0)
        
        # If split not available, try to get from production/pre-production
        if ec2_backup_monthly == 0 and eks_backup_monthly == 0:
            prod_backup = backup_costs.get('production', {}).get('monthly_cost', 0)
            preprod_backup = backup_costs.get('pre_production', {}).get('monthly_cost', 0)
            
            if prod_backup > 0 or preprod_backup > 0:
                # Show split by environment
                backup_components = [
                    ('Production Backups', f'{backup_costs.get("production", {}).get("vm_count", 0)} VMs', prod_backup),
                    ('Pre-Production Backups', f'{backup_costs.get("pre_production", {}).get("vm_count", 0)} VMs', preprod_backup),
                ]
            else:
                # Show total only
                backup_components = [
                    ('AWS Backup', f'All {len(ec2_vms) + len(eks_vms)} VMs', backup_monthly),
                ]
        else:
            # Show EC2 and EKS split
            backup_components = [
                ('EC2 Backup', f'{len(ec2_vms)} VMs', ec2_backup_monthly),
                ('EKS Backup', f'{len(eks_vms)} persistent volumes', eks_backup_monthly),
            ]
    else:
        # No backup costs provided
        backup_components = [
            ('AWS Backup', 'Not included', 0),
        ]
    
    for component, details, cost in backup_components:
        ws[f'A{row}'] = component
        ws[f'B{row}'] = details
        ws[f'D{row}'] = cost
        ws[f'D{row}'].number_format = '$#,##0.00'
        
        for col in range(1, 5):
            ws.cell(row, col).border = border
            ws.cell(row, col).fill = backup_fill
        
        row += 1
    
    # Backup Subtotal
    ws[f'A{row}'] = 'Backup Subtotal'
    ws[f'A{row}'].font = Font(bold=True)
    # FIX: Use EC2 + EKS backup sum, not total_monthly (which includes all VMs)
    backup_subtotal = ec2_backup_monthly + eks_backup_monthly
    ws[f'D{row}'] = backup_subtotal
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    for col in range(1, 5):
        ws.cell(row, col).border = border
        ws.cell(row, col).fill = section_fill
    
    # Section 4: Total Hybrid Cost (including backup)
    row += 2
    total_monthly = ec2_total + eks_total + backup_subtotal
    total_annual = total_monthly * 12
    total_3yr = total_monthly * 36
    
    totals = [
        ('TOTAL HYBRID MONTHLY COST', total_monthly),
        ('TOTAL HYBRID ANNUAL COST (ARR)', total_annual),
        ('TOTAL HYBRID 3-YEAR COST', total_3yr),
    ]
    
    for label, value in totals:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'] = value
        ws[f'D{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:C{row}')
        
        for col in range(1, 5):
            ws.cell(row, col).border = border
            ws.cell(row, col).fill = total_fill
        
        row += 1
    
    # Summary statistics
    row += 2
    ws[f'A{row}'] = 'Migration Summary'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    summary_stats = [
        ('Total VMs', len(ec2_vms) + len(eks_vms)),
        ('VMs migrating to EC2', len(ec2_vms)),
        ('VMs containerized on EKS', len(eks_vms)),
        ('EC2 percentage', f"{len(ec2_vms) / (len(ec2_vms) + len(eks_vms)) * 100:.1f}%"),
        ('EKS percentage', f"{len(eks_vms) / (len(ec2_vms) + len(eks_vms)) * 100:.1f}%"),
        ('EC2 cost share', f"${ec2_total:,.2f} ({ec2_total / total_monthly * 100:.1f}%)"),
        ('EKS cost share', f"${eks_total:,.2f} ({eks_total / total_monthly * 100:.1f}%)"),
        ('Backup cost share', f"${backup_monthly:,.2f} ({backup_monthly / total_monthly * 100:.1f}%)"),
    ]
    
    for label, value in summary_stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws.merge_cells(f'B{row}:D{row}')
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    return ws


def create_migration_strategy_summary_tab(workbook, strategy_recommendation, mra_scores=None):
    """
    Create Migration Strategy Summary tab
    
    Executive summary of EKS vs EC2 recommendation
    
    Args:
        workbook: openpyxl Workbook object
        strategy_recommendation: Dict from recommend_eks_strategy()
        mra_scores: Optional dict with MRA scores
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = workbook.create_sheet('Migration Strategy Summary')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'EKS Migration Strategy - Executive Summary'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    
    # Recommended Strategy
    ws['A3'] = 'Recommended Strategy'
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = section_fill
    ws.merge_cells('A3:C3')
    
    strategy = strategy_recommendation.get('recommended_strategy', 'hybrid')
    confidence = strategy_recommendation.get('confidence', 'medium')
    
    ws['A4'] = 'Strategy'
    ws['B4'] = strategy.upper()
    ws['A4'].font = Font(bold=True)
    ws['B4'].font = Font(bold=True, size=12, color='0000FF')
    
    ws['A5'] = 'Confidence Level'
    ws['B5'] = confidence.upper()
    ws['A5'].font = Font(bold=True)
    
    # Rationale
    row = 7
    ws[f'A{row}'] = 'Rationale'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    rationale = strategy_recommendation.get('rationale', 'No rationale provided')
    # Convert list to string if needed
    if isinstance(rationale, list):
        rationale = '\n'.join(f"• {item}" for item in rationale)
    ws[f'A{row}'] = rationale
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(f'A{row}:C{row}')
    ws.row_dimensions[row].height = 60
    
    # MRA Scores (if available)
    if mra_scores:
        row += 2
        ws[f'A{row}'] = 'MRA Assessment Scores'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = 'Factor'
        ws[f'B{row}'] = 'Score (1-5)'
        ws[f'C{row}'] = 'Assessment'
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        mra_data = [
            ('Cloud Readiness', mra_scores.get('cloud_readiness', 3)),
            ('Container Expertise', mra_scores.get('container_expertise', 2)),
            ('DevOps Maturity', mra_scores.get('devops_maturity', 2)),
            ('Change Readiness', mra_scores.get('change_readiness', 3)),
        ]
        
        row += 1
        for factor, score in mra_data:
            assessment = 'High' if score >= 4 else 'Medium' if score >= 3 else 'Low'
            ws[f'A{row}'] = factor
            ws[f'B{row}'] = score
            ws[f'C{row}'] = assessment
            
            for col in range(1, 4):
                ws.cell(row, col).border = border
            row += 1
    
    # Timeline
    row += 1
    ws[f'A{row}'] = 'Estimated Timeline'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    timeline = strategy_recommendation.get('timeline_months', 12)
    ws[f'A{row}'] = 'Migration Duration'
    ws[f'B{row}'] = f'{timeline} months'
    ws[f'A{row}'].font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    
    return ws


def create_eks_roi_analysis_tab(workbook, ec2_costs, eks_costs, vm_count, vm_categorization=None):
    """
    Create EKS ROI Analysis tab
    
    Shows ROI calculations for EKS VMs only
    
    Args:
        workbook: openpyxl Workbook object
        ec2_costs: Dict with EC2 pricing (for EKS VMs only)
        eks_costs: Dict with EKS pricing
        vm_count: Number of VMs planned for EKS migration
        vm_categorization: Optional dict from categorize_vms_for_eks() for accurate VM count
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = workbook.create_sheet('EKS ROI Analysis')
    
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    savings_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    cost_increase_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f'EKS ROI Analysis ({vm_count} VMs)'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Cost breakdown
    ec2_compute = ec2_costs.get('compute_monthly', 0)
    ec2_storage = ec2_costs.get('storage_monthly', 0)
    
    eks_compute = eks_costs.get('compute_monthly', 0)
    eks_storage = eks_costs.get('storage_monthly', 0)
    eks_control_plane = eks_costs.get('control_plane_monthly', 0)
    eks_networking = eks_costs.get('networking_monthly', 0)
    eks_monitoring = eks_costs.get('monitoring_monthly', 0)
    
    # Equal costs for fair comparison
    data_transfer_monthly = 50
    ec2_networking_monthly = eks_networking  # Use same networking cost (10% of compute)
    ec2_monitoring_monthly = eks_monitoring  # Use same monitoring cost
    
    ec2_monthly = ec2_compute + ec2_storage + data_transfer_monthly + ec2_networking_monthly + ec2_monitoring_monthly
    eks_monthly = eks_compute + eks_storage + eks_control_plane + eks_networking + eks_monitoring + data_transfer_monthly
    
    # Cost Comparison
    ws['A3'] = 'Cost Comparison (EKS VMs Only)'
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = section_fill
    ws.merge_cells('A3:D3')
    
    difference = eks_monthly - ec2_monthly  # Positive = EKS costs more, Negative = EKS saves
    annual_difference = difference * 12
    three_year_difference = difference * 36
    
    cost_data = [
        ('EC2-Only Monthly Cost (3yr RI)', ec2_monthly, ''),
        ('EKS Hybrid Monthly Cost (3yr RI)', eks_monthly, ''),
        ('Monthly Difference', difference, 'EKS costs more' if difference > 0 else 'EKS saves'),
        ('Annual Difference (Year 1)', annual_difference, ''),
        ('3-Year Difference', three_year_difference, ''),
    ]
    
    row = 4
    for label, value, note in cost_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'] = note
        ws[f'A{row}'].font = Font(bold=True)
        
        if 'Difference' in label:
            ws[f'B{row}'].fill = savings_fill if value < 0 else cost_increase_fill
        
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'C{row}'].border = border
        row += 1
    
    # Efficiency Metrics
    row += 1
    ws[f'A{row}'] = 'Efficiency Metrics'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    cost_per_vm_ec2 = ec2_monthly / vm_count if vm_count > 0 else 0
    cost_per_vm_eks = eks_monthly / vm_count if vm_count > 0 else 0
    difference_per_vm = cost_per_vm_eks - cost_per_vm_ec2
    
    # Get actual EKS VM count from summary
    if vm_categorization:
        summary = vm_categorization.get('summary', {})
        eks_vm_count = summary.get('eks_total', vm_count)  # Use eks_total from categorization
    else:
        eks_vm_count = vm_count  # Fallback to passed vm_count
    
    efficiency_data = [
        ('Total VMs (EKS Migration)', eks_vm_count, ''),
        ('Cost per VM (EC2 3yr RI)', cost_per_vm_ec2, '/month'),
        ('Cost per VM (EKS 3yr RI)', cost_per_vm_eks, '/month'),
        ('Difference per VM', difference_per_vm, '/month'),
    ]
    
    for label, value, unit in efficiency_data:
        ws[f'A{row}'] = label
        if isinstance(value, (int, float)) and 'Total VMs' not in label:
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '$#,##0.00'
            ws[f'C{row}'] = unit
            if 'Difference' in label:
                ws[f'B{row}'].fill = savings_fill if value < 0 else cost_increase_fill
        else:
            ws[f'B{row}'] = value
        
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'C{row}'].border = border
        row += 1
    
    # Strategic Value Note
    row += 1
    ws[f'A{row}'] = 'Strategic Value Beyond Cost'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    strategic_notes = [
        '• Container-native architecture enables microservices and modern DevOps practices',
        '• Auto-scaling and resource efficiency improve application performance',
        '• Kubernetes ecosystem provides access to cloud-native tools and services',
        '• Reduced operational overhead through managed EKS control plane',
        '• Future-proof infrastructure aligned with industry best practices',
    ]
    
    for note in strategic_notes:
        ws[f'A{row}'] = note
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    
    return ws


def create_backup_cost_details_tab(workbook, vm_categorization, backup_costs, region='us-east-1'):
    """
    Create detailed backup cost breakdown tab for hybrid scenario
    Shows EC2 VM backup and EKS PV backup calculations with notes
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    ws = workbook.create_sheet('Backup Cost Details', 1)  # Insert as second tab
    
    # Header styling
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_font = Font(bold=True, size=11)
    note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = 'Backup Cost Details - Hybrid Migration Scenario'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f'Detailed breakdown of backup costs for EC2 VMs and EKS persistent volumes'
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Get VM lists
    ec2_vms = vm_categorization.get('ec2', [])
    eks_vms = vm_categorization.get('eks_linux', []) + vm_categorization.get('eks_windows', [])
    
    # Get backup costs
    ec2_backup_monthly = backup_costs.get('ec2_backup_monthly', 0) if backup_costs else 0
    eks_backup_monthly = backup_costs.get('eks_backup_monthly', 0) if backup_costs else 0
    total_backup = ec2_backup_monthly + eks_backup_monthly
    
    row = 4
    
    # Summary Section
    ws[f'A{row}'] = 'BACKUP COST SUMMARY'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    summary_data = [
        ('Component', 'VM/PV Count', 'Storage (GB)', 'Monthly Cost', 'Annual Cost'),
        ('EC2 VM Backups', len(ec2_vms), '', f'${ec2_backup_monthly:,.2f}', f'${ec2_backup_monthly * 12:,.2f}'),
        ('EKS Persistent Volume Backups', len(eks_vms), '', f'${eks_backup_monthly:,.2f}', f'${eks_backup_monthly * 12:,.2f}'),
        ('TOTAL HYBRID BACKUP COST', len(ec2_vms) + len(eks_vms), '', f'${total_backup:,.2f}', f'${total_backup * 12:,.2f}'),
    ]
    
    for data_row in summary_data:
        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row, col_idx, value)
            cell.border = border
            if row == 5:  # Header row
                cell.fill = header_fill
                cell.font = header_font
            elif 'TOTAL' in str(data_row[0]):
                cell.font = Font(bold=True)
                cell.fill = section_fill
        row += 1
    
    row += 1
    
    # EC2 Backup Details Section
    ws[f'A{row}'] = 'EC2 VM BACKUP DETAILS'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # EC2 VM list header
    ec2_headers = ['VM Name', 'vCPU', 'Memory (GB)', 'Storage (GB)', 'Environment']
    for col_idx, header in enumerate(ec2_headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    row += 1
    
    # EC2 VM list
    for vm in ec2_vms:
        vm_data = [
            vm.get('name', vm.get('vm_name', 'Unknown')),
            vm.get('vcpu', 0),
            vm.get('memory', vm.get('memory_gb', 0)),
            vm.get('storage', vm.get('storage_gb', 0)),
            vm.get('environment', 'Production')
        ]
        for col_idx, value in enumerate(vm_data, 1):
            cell = ws.cell(row, col_idx, value)
            cell.border = border
        row += 1
    
    row += 1
    
    # EKS Backup Details Section
    ws[f'A{row}'] = 'EKS PERSISTENT VOLUME BACKUP DETAILS'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # EKS VM list header
    eks_headers = ['VM Name (Containerized)', 'vCPU', 'Memory (GB)', 'Storage (GB)', 'Stateful']
    for col_idx, header in enumerate(eks_headers, 1):
        cell = ws.cell(row, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    row += 1
    
    # EKS VM list
    for vm in eks_vms:
        vm_data = [
            vm.get('name', vm.get('vm_name', 'Unknown')),
            vm.get('vcpu', 0),
            vm.get('memory', vm.get('memory_gb', 0)),
            vm.get('storage', vm.get('storage_gb', 0)),
            'Yes' if vm.get('storage', vm.get('storage_gb', 0)) > 0 else 'No'
        ]
        for col_idx, value in enumerate(vm_data, 1):
            cell = ws.cell(row, col_idx, value)
            cell.border = border
        row += 1
    
    row += 2
    
    # Important Notes Section
    ws[f'A{row}'] = 'IMPORTANT NOTES ON BACKUP COSTS'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = note_fill
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    notes = [
        '1. AWS Backup Service: Both EC2 and EKS backups use AWS Backup service with the same pricing model',
        '2. EC2 VM Backups: Full VM snapshots including all EBS volumes attached to each EC2 instance',
        '3. EKS Persistent Volume Backups: Only stateful workloads require persistent volumes and backups',
        '4. Backup Frequency: Daily backups with 30-day retention (warm storage) + 90-day cold storage',
        '5. Incremental Snapshots: AWS uses incremental snapshots, only changed blocks are stored',
        '6. Deduplication: No cross-volume deduplication (deduplication_ratio = 1.0)',
        '7. Compression: AWS provides automatic compression for snapshots',
        '8. Cost Factors: Backup cost depends on total storage size and retention period',
        '',
        'WHY HYBRID BACKUP COSTS DIFFER FROM ALL-EC2:',
        '• Hybrid (30 EC2 + 21 EKS): Lower backup cost because fewer VMs on EC2',
        '• All-EC2 (51 VMs): Higher backup cost because all VMs require full snapshots',
        '• EKS PV backups are typically smaller than full VM backups',
        '',
        'BACKUP COST CALCULATION:',
        f'• EC2 Backup ({len(ec2_vms)} VMs): ${ec2_backup_monthly:,.2f}/month',
        f'• EKS Backup ({len(eks_vms)} PVs): ${eks_backup_monthly:,.2f}/month',
        f'• Total: ${total_backup:,.2f}/month',
        '',
        'For detailed backup pricing methodology, see AWS Backup Calculator Guide'
    ]
    
    for note in notes:
        ws[f'A{row}'] = note
        ws[f'A{row}'].fill = note_fill
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20


def export_eks_analysis_to_excel(vm_categorization, cluster_config, ec2_costs, eks_costs, 
                                  strategy_recommendation, mra_scores=None, backup_costs=None,
                                  output_filename='eks_analysis.xlsx', region='us-east-1'):
    """
    Main function to export complete EKS analysis to Excel with 9 tabs
    
    Tabs:
    1. Hybrid Migration Summary - Complete cost breakdown (EC2 + EKS + Backup)
    2. Backup Cost Details - Detailed backup breakdown for EC2 VMs and EKS PVs (NEW)
    3. Migration Strategy Summary - Strategy recommendation and rationale
    4. EKS VM Details - Individual VM costs for EKS migration
    5. EKS Capacity Mix - Reserved/On-Demand/Spot breakdown
    6. EKS VM Categorization - VM sizing and categorization
    7. EKS Cluster Design - Worker node configuration
    8. EC2 vs EKS Comparison - Cost comparison for EKS VMs
    9. EKS ROI Analysis - Return on investment analysis
    
    Args:
        vm_categorization: Dict from categorize_vms_for_eks()
        cluster_config: Dict from calculate_eks_cluster_size()
        ec2_costs: Dict with EC2 pricing
        eks_costs: Dict with EKS pricing
        strategy_recommendation: Dict from recommend_eks_strategy()
        mra_scores: Optional dict with MRA scores
        backup_costs: Optional dict with backup costs (from calculate_backup_costs)
        output_filename: Output Excel filename
        region: AWS region
    
    Returns:
        Path to generated Excel file
    """
    from openpyxl import Workbook
    from agents.pricing.aws_pricing_calculator import AWSPricingCalculator
    import logging
    
    logger = logging.getLogger('AgentWorkflow')
    
    output_path = get_case_output_path(output_filename)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get EKS VMs
    strategy = strategy_recommendation.get('recommended_strategy', 'hybrid')
    summary = vm_categorization.get('summary', {})
    eks_vms = vm_categorization.get('eks_linux', []) + vm_categorization.get('eks_windows', [])
    vm_count = len(eks_vms)
    
    # CRITICAL: Read EC2 costs from pricing Excel FIRST before using them
    # This must happen before ec2_costs_all_vms is created
    ec2_cost_map = {}
    backup_costs_from_excel = None
    
    # FIX: Try to get ec2_cost_map from backup_costs parameter first (passed from business case)
    if backup_costs and 'ec2_cost_map' in backup_costs:
        ec2_cost_map = backup_costs['ec2_cost_map']
        logger.info(f"✓ Using EC2 cost map from backup_costs parameter ({len(ec2_cost_map)} VMs)")
    
    # If ec2_cost_map not provided or empty, try to read from Excel file
    if not ec2_cost_map:
        try:
            # Find the pricing Excel file
            import glob
            from agents.utils.project_context import get_case_id
            from agents.config.config import output_folder_dir_path
            case_id = get_case_id()
            
            pricing_excel = None
            df_ec2 = None
            
            if case_id:
                # Use absolute path from config
                case_output_dir = os.path.join(output_folder_dir_path, case_id)
                
                logger.info(f"Looking for pricing Excel in: {case_output_dir}")
                
                # Try IT Inventory first
                excel_files = glob.glob(os.path.join(case_output_dir, 'it_inventory_aws_pricing_*.xlsx'))
                if not excel_files:
                    # Try RVTools
                    rvtools_path = os.path.join(case_output_dir, 'vm_to_ec2_mapping.xlsx')
                    if os.path.exists(rvtools_path):
                        excel_files = [rvtools_path]
                if excel_files:
                    pricing_excel = excel_files[0]
                    logger.info(f"Found pricing Excel: {pricing_excel}")
            
            if pricing_excel and os.path.exists(pricing_excel):
                # Determine sheet name based on file type
                # IT Inventory uses: EC2_Option1_Instance_SP
                # RVTools uses: EC2 Details - Option 1
                sheet_name = 'EC2_Option1_Instance_SP' if 'it_inventory' in os.path.basename(pricing_excel) else 'EC2 Details - Option 1'
                
                # Read EC2 Details sheet
                df_ec2 = pd.read_excel(pricing_excel, sheet_name=sheet_name)
                logger.info(f"✓ Reading actual EC2 costs from {os.path.basename(pricing_excel)} (sheet: {sheet_name})")
            
            # Create a mapping of VM name to EC2 cost (only if df_ec2 was successfully loaded)
            if df_ec2 is not None:
                for idx, row in df_ec2.iterrows():
                    # IT Inventory uses 'Hostname', RVTools uses 'VM Name'
                    vm_name = row.get('Hostname', row.get('VM Name', ''))
                    
                    # IT Inventory uses 'Monthly Cost', RVTools uses 'Total Monthly Cost ($)'
                    total_cost = row.get('Monthly Cost', row.get('Total Monthly Cost ($)', 0))
                    
                    if vm_name and total_cost:
                        # Remove $ and commas from cost if it's a string
                        if isinstance(total_cost, str):
                            total_cost = total_cost.replace('$', '').replace(',', '')
                        
                        ec2_cost_map[vm_name] = float(total_cost)
                        
                        # Also add variant with hyphen (Server1 -> Server-1)
                        # This handles naming mismatches between Excel and categorization
                        if vm_name and vm_name[0].isalpha():
                            # Find where digits start
                            for i, char in enumerate(vm_name):
                                if char.isdigit():
                                    vm_name_with_hyphen = vm_name[:i] + '-' + vm_name[i:]
                                    ec2_cost_map[vm_name_with_hyphen] = float(total_cost)
                                    break
                
                logger.info(f"✓ Loaded EC2 costs for {len(ec2_cost_map)} VMs from pricing sheet")
            
            # Also read backup costs from Pricing Comparison sheet
            try:
                sheet_name = 'Pricing_Comparison' if 'it_inventory' in os.path.basename(pricing_excel) else 'Pricing Comparison'
                df_pricing = pd.read_excel(pricing_excel, sheet_name=sheet_name)
                
                # Find backup cost row
                for idx, row in df_pricing.iterrows():
                    if 'backup monthly cost' in str(row.get('Metric', '')).lower():
                        backup_monthly = float(str(row.get('Value', '0')).replace('$', '').replace(',', ''))
                        if backup_monthly > 0:
                            backup_costs_from_excel = {
                                'total_monthly': backup_monthly,
                                'total_annual': backup_monthly * 12
                            }
                            logger.info(f"✓ Loaded backup costs: ${backup_monthly:,.2f}/month")
                            break
            except Exception as e:
                    logger.warning(f"Could not read backup costs from Excel: {e}")
            else:
                logger.warning(f"⚠ Could not find pricing Excel (case_id={case_id}, pricing_excel={pricing_excel})")
        except Exception as e:
            logger.error(f"⚠ Could not read EC2 costs from Excel: {e}")
            import traceback
            logger.error(traceback.format_exc())
    
    # Use backup costs from parameter if provided, otherwise use from Excel
    if backup_costs is None:
        backup_costs = backup_costs_from_excel
    
    # Calculate per-VM costs for the new VM Details tab
    # ec2_cost_map was already created at the beginning of this function
    calculator = AWSPricingCalculator(region=region)
    ec2_costs_per_vm = {}
    eks_costs_per_vm = {}
    
    # DEBUG: Log EKS costs to diagnose discrepancy
    logger.info("="*80)
    logger.info("EKS COST BREAKDOWN DEBUG (export_eks_analysis_to_excel)")
    logger.info("="*80)
    logger.info(f"  monthly_avg:           ${eks_costs.get('monthly_avg', 0):,.2f}")
    logger.info(f"  compute_monthly:       ${eks_costs.get('compute_monthly', 0):,.2f}")
    logger.info(f"  storage_monthly:       ${eks_costs.get('storage_monthly', 0):,.2f}")
    logger.info(f"  control_plane_monthly: ${eks_costs.get('control_plane_monthly', 0):,.2f}")
    logger.info(f"  networking_monthly:    ${eks_costs.get('networking_monthly', 0):,.2f}")
    logger.info(f"  monitoring_monthly:    ${eks_costs.get('monitoring_monthly', 0):,.2f}")
    
    component_sum = (
        eks_costs.get('compute_monthly', 0) +
        eks_costs.get('storage_monthly', 0) +
        eks_costs.get('control_plane_monthly', 0) +
        eks_costs.get('networking_monthly', 0) +
        eks_costs.get('monitoring_monthly', 0)
    )
    logger.info(f"  Sum of components:     ${component_sum:,.2f}")
    logger.info(f"  Discrepancy:           ${eks_costs.get('monthly_avg', 0) - component_sum:,.2f}")
    logger.info("="*80)
    
    # FIX: Use component sum instead of monthly_avg to ensure consistency
    # This ensures VM Details tab matches the Comparison tab breakdown
    total_eks_monthly = component_sum
    total_vcpu = sum(vm.get('vcpu', 0) for vm in eks_vms)
    
    for vm in eks_vms:
        vm_name = vm.get('name', vm.get('vm_name', ''))
        vcpu = vm.get('vcpu', 0)
        memory_gb = vm.get('memory', vm.get('memory_gb', 0))
        storage_gb = vm.get('storage_gb', 100)
        os_type = 'Windows' if 'windows' in vm.get('os', '').lower() else 'Linux'
        
        # Use actual EC2 cost from pricing sheet if available
        if vm_name in ec2_cost_map:
            ec2_costs_per_vm[vm_name] = ec2_cost_map[vm_name]
        else:
            # Fallback: Calculate EC2 cost for this VM
            instance_type = calculator.map_vm_to_instance_type(vcpu, memory_gb, os_type)
            try:
                hourly_rate = calculator.get_ec2_price_by_term(
                    instance_type=instance_type,
                    os_type=os_type,
                    region=region,
                    term='3yr',
                    purchase_option='No Upfront'
                )
                ec2_compute_monthly = hourly_rate * 730
            except:
                ec2_compute_monthly = vcpu * 30  # Fallback estimate
            
            # Add storage
            ebs_rate = calculator.get_ebs_gp3_price(region)
            storage_monthly = storage_gb * ebs_rate
            
            ec2_costs_per_vm[vm_name] = ec2_compute_monthly + storage_monthly
        
        # Calculate EKS cost share for this VM (proportional to vCPU)
        vm_share = vcpu / total_vcpu if total_vcpu > 0 else 0
        eks_costs_per_vm[vm_name] = total_eks_monthly * vm_share
    
    # Get ALL VM costs for Hybrid Migration Summary tab
    # CRITICAL: Use actual costs from EC2 Details sheet (includes right-sizing)
    # DO NOT recalculate - that would miss right-sizing optimizations
    ec2_costs_all_vms = {}
    
    # Add EC2 VMs (staying on EC2) - use actual costs from pricing sheet
    for vm in vm_categorization.get('ec2', []):
        vm_name = vm.get('name', vm.get('vm_name', ''))
        if vm_name in ec2_cost_map:
            ec2_costs_all_vms[vm_name] = ec2_cost_map[vm_name]
            logger.info(f"  EC2 VM '{vm_name}': ${ec2_cost_map[vm_name]:,.2f} (from pricing sheet)")
        else:
            logger.warning(f"  EC2 VM '{vm_name}' not found in ec2_cost_map, skipping")
            # DO NOT calculate - if VM is not in the pricing sheet, something is wrong
            # Better to show 0 and investigate than to show wrong costs
            ec2_costs_all_vms[vm_name] = 0
    
    # Create all tabs (added Backup Cost Details tab)
    create_hybrid_migration_summary_tab(wb, vm_categorization, ec2_costs_all_vms, eks_costs, backup_costs, region)
    create_backup_cost_details_tab(wb, vm_categorization, backup_costs, region)  # NEW: Detailed backup breakdown
    create_migration_strategy_summary_tab(wb, strategy_recommendation, mra_scores)
    create_eks_vm_details_tab(wb, eks_vms, ec2_costs_per_vm, eks_costs_per_vm, region)
    create_eks_capacity_mix_tab(wb, cluster_config, eks_costs, region)
    create_eks_sizing_tab(wb, vm_categorization, strategy)
    create_eks_cluster_design_tab(wb, cluster_config, region)
    create_ec2_vs_eks_comparison_tab(wb, ec2_costs, eks_costs, vm_categorization, cluster_config)
    create_eks_roi_analysis_tab(wb, ec2_costs, eks_costs, vm_count, vm_categorization)
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ EKS analysis Excel created: {output_path}")
    
    return output_path
